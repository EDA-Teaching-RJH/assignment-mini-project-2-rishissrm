import csv
import os 
import random 
import statistics
import re
from datetime import datetime 
from typing import List 
import requests
from openpyxl import workbook

#banking system

class person:
    def __init__(self,name,age):
       self.name=name
       self.age=age

#customer input for the data collection
#these belong to this specific student object

class customer(person):
    def __init__(self,customer_id,name,age,email,phone,balance=0.0):
        super().__init__(name,age)
        self.customer_id=customer_id
        self.phone=phone
        self.balance=balance
        self.email=email


    def deposit(self,amount):
        if amount<=0:
            raise ValueError("deposit amount must be greater than 0.")
        self.balnce+= amount

        #withdrw money from the customers account
        def withdraw(self,amount):
            if amount <=0:
                raise ValueError("withdrawal amount must be greater than 0.")
            if amount > self.balance:
                raise ValueError("Insuffient funds.")
            self.balance -=amount

    #coverting customer data into list to be saved in cvs
    def to_list(self):
        return[
             self.customer_id,
             self.name,
             self.age,
             self.email,
             self.phone,
             round(self.balance,2)
        ]


        #transaction class stores one transaction record.
        class transaction:
            def __init__(self,transaction_id,customer_id,transaction_type,amount,date):
               self.transaction_id=transaction_id
               self.customer_id=customer_id
               self.transaction_type=transaction_type
               self.amount=amount
               self.date=date
#Convert transaction data into a list for CSV or excel
        def to_list(self):
            return [
                self.transaction_id,
                self.customer_id,
                self.transaction_type,
                round(self.amount,2),
                self.date
            ]

#Bank class is the main part of the program
#It stores all the customers and their transations and performs all operations
class Bank:
    def __init__(self):
        self.customers: List[Customer] = []
        self.transations: List[Transactions] = []

    #Check whether the name contrains only letters and spaces.
    def is_valid_name(self,name):
        pattern = r"^[A-Za-z ]{2,30}$"
        return bool(re.match(pattern,name))
    
    #Check whether the email is in a valid format
    def is_valid_email(self,email):
        pattern = r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[A-Za-z]{2,}$"
        return bool(re.match(pattern,email))

    #Check whether the phone number has 10 or 11 digits.
    def is_valid_phone(self,phone):
        pattern = r"^[0-9]{10,11}$"
        return bool(re.match(pattern,phone))

    #Add a new customer after validating the input.
    def add_customer(self,name,age,email,phone,opening_balance=0.0):
        if not self.is_valid_name(name):
            raise ValueError("Invalid name, Use Letters and spaces only.")
        
        if not self.is_valid_email(email):
            raise ValueError("Invalid email format.")
        
        if not self.is_valid_phone(phone):
            raise ValueError("Invalid phone number. Use 10 or 11 digits.")

        customer_id = len(self.customer) +1
        customer = Customer(customer_id,name, age,email,phone,opening_balance)
        self.customers.append(customer)
        return Customer

    #Find a customer by their customer ID.
    def find_customer(self,customer_id):
        for customer in sef.customers:
            if customer.customer_id == customer_id:
                return customer
        raise ValueError("Customer not found.")

    #Save one transation into the transaction list.
    def record_transaction(self,customer_id,transaction_type,amount):
        transaction_id =len(self.transations) +1
        date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        transaction = transaction(transaction_id,customer_id,transaction_type,amount,date)
        self.transations.append(transaction)

    #Deposit money into a specific customer's account:
    def deposit_to_customer(self,customer_id,amount):
        customer = self.find_customer(customer_id)
        customer.deposit(amount)
        self.record_transaction(customer_id,"Deposit",amount)

    #Withdraw money from a specific customer's account.
    def withdraw_from_customer(self,customer_id,amount):
        customer = self.find_customer(customer_id)
        customer.withdraw(amount)
        self.record_transaction(customer_id,"Withdrawal",amount)

    #Transfer money from one customer to another.
    def transfer(self,from_customer_id,to_customer_id,amount):
        sender = self.find_customer(from_customer_id)
        receiver = self.find_customer(to_customer_id)

        sender.withdraw(amount)
        receiver.deposit(amount)

        self.record_transaction(from_customer_id,"Transfer Out",amount)
        self.record_transaction(to_customer_id,"Transfer In",amount)

    # Create sample customers and sample transactions automatically.
    # This is useful for testing the program quickly.
    def generate_sample_data(self, num_customers=5, num_transactions=10):
        names = ["Alice", "Bob", "Charlie", "Diana", "Eva", "Farhan", "George", "Hana", "Imran", "Sophie"]

        for i in range(num_customers):
            name = random.choice(names)
            age = random.randint(18, 60)
            email = name.lower() + str(i + 1) + "@gmail.com"
            phone = "07" + str(random.randint(100000000, 999999999))
            balance = round(random.uniform(100, 5000), 2)
            self.add_customer(name, age, email, phone, balance)

        for _ in range(num_transactions):
            customer = random.choice(self.customers)
            action = random.choice(["deposit", "withdraw"])
            amount = round(random.uniform(10, 400), 2)

            try:
                if action == "deposit":
                    self.deposit_to_customer(customer.customer_id, amount)
                else:
                    self.withdraw_from_customer(customer.customer_id, amount)
            except ValueError:
                # Ignore invalid random withdrawals, for example if balance is too low.
                pass

# Calculate useful transaction statistics.
    def transaction_statistics(self):
        amounts = [t.amount for t in self.transactions]

        if not amounts:
            return {
                "count": 0,
                "total": 0,
                "mean": 0,
                "median": 0,
                "max": 0,
                "min": 0
            }

        return {
            "count": len(amounts),
            "total": round(sum(amounts), 2),
            "mean": round(statistics.mean(amounts), 2),
            "median": round(statistics.median(amounts), 2),
            "max": round(max(amounts), 2),
            "min": round(min(amounts), 2)
        }

# Save all customers into a CSV file.
    def save_customers_to_csv(self, filename="customers.csv"):
        with open(filename, "w", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow(["CustomerID", "Name", "Age", "Email", "Phone", "BalanceGBP"])
            for customer in self.customers:
                writer.writerow(customer.to_list())

# Load customer data from a CSV file.
    def load_customers_from_csv(self, filename="customers.csv"):
        if not os.path.exists(filename):
            raise FileNotFoundError(f"{filename} not found.")

        self.customers.clear()

        with open(filename, "r", newline="", encoding="utf-8") as file:
            reader = csv.DictReader(file)
            for row in reader:
                customer = Customer(
                    int(row["CustomerID"]),
                    row["Name"],
                    int(row["Age"]),
                    row["Email"],
                    row["Phone"],
                    float(row["BalanceGBP"])
                )
                self.customers.append(customer)

 # Save all transactions into a CSV file.
    def save_transactions_to_csv(self, filename="transactions.csv"):
        with open(filename, "w", newline="", encoding="utf-8") as file:
            writer = csv.writer(file)
            writer.writerow(["TransactionID", "CustomerID", "Type", "AmountGBP", "Date"])
            for transaction in self.transactions:
                writer.writerow(transaction.to_list())

# Load transactions from a CSV file.
    def load_transactions_from_csv(self, filename="transactions.csv"):
        if not os.path.exists(filename):
            raise FileNotFoundError(f"{filename} not found.")

        self.transactions.clear()

        with open(filename, "r", newline="", encoding="utf-8") as file:
            reader = csv.DictReader(file)
            for row in reader:
                transaction = Transaction(
                    int(row["TransactionID"]),
                    int(row["CustomerID"]),
                    row["Type"],
                    float(row["AmountGBP"]),
                    row["Date"]
                )
                self.transactions.append(transaction)

# Save a simple summary report to a TXT file.
    def save_summary_to_txt(self, filename="summary_report.txt"):
        stats = self.transaction_statistics()

        with open(filename, "w", encoding="utf-8") as file:
            file.write("BANK SUMMARY REPORT\n")
            file.write("=" * 35 + "\n")
            file.write(f"Number of customers: {len(self.customers)}\n")
            file.write(f"Number of transactions: {stats['count']}\n")
            file.write(f"Total transaction amount: GBP {stats['total']}\n")
            file.write(f"Average transaction: GBP {stats['mean']}\n")
            file.write(f"Median transaction: GBP {stats['median']}\n")
            file.write(f"Maximum transaction: GBP {stats['max']}\n")
            file.write(f"Minimum transaction: GBP {stats['min']}\n")

# Export customers, transactions, and summary to an Excel file.
    def export_to_excel(self, filename="bank_report.xlsx"):
        workbook = Workbook()

        # First worksheet for customer details.
        ws1 = workbook.active
        ws1.title = "Customers"
        ws1.append(["CustomerID", "Name", "Age", "Email", "Phone", "BalanceGBP"])
        for customer in self.customers:
            ws1.append(customer.to_list())

        # Second worksheet for transaction details.
        ws2 = workbook.create_sheet(title="Transactions")
        ws2.append(["TransactionID", "CustomerID", "Type", "AmountGBP", "Date"])
        for transaction in self.transactions:
            ws2.append(transaction.to_list())

        # Third worksheet for summary statistics.
        ws3 = workbook.create_sheet(title="Summary")
        stats = self.transaction_statistics()
        ws3.append(["Metric", "Value"])
        for key, value in stats.items():
            ws3.append([key, value])

        workbook.save(filename)

# Use an API to get the next UK bank holiday.
    def get_next_bank_holiday(self):
        url = "https://www.gov.uk/bank-holidays.json"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        data = response.json()

        holidays = data["england-and-wales"]["events"]
        today = datetime.now().date()

        for holiday in holidays:
            holiday_date = datetime.strptime(holiday["date"], "%Y-%m-%d").date()
            if holiday_date >= today:
                return f"{holiday['title']} on {holiday['date']}"

        return "No upcoming bank holiday found."


 # Display the menu and allow the user to choose options.
    def menu(self):
        while True:
            print("\n--- BANK MENU ---")
            print("1. Add customer")
            print("2. Deposit money")
            print("3. Withdraw money")
            print("4. Transfer money")
            print("5. Show statistics")
            print("6. Generate sample data")
            print("7. Save data to CSV/TXT")
            print("8. Export report to Excel")
            print("9. Show next UK bank holiday")
            print("10. Exit")

            choice = input("Enter your choice: ")

            try:
                if choice == "1":
                    name = input("Enter customer name: ")
                    age = int(input("Enter age: "))
                    email = input("Enter email: ")
                    phone = input("Enter phone number: ")
                    balance = float(input("Enter opening balance in GBP: "))
                    customer = self.add_customer(name, age, email, phone, balance)
                    print("Customer added with ID", customer.customer_id)

                elif choice == "2":
                    customer_id = int(input("Enter customer ID: "))
                    amount = float(input("Enter deposit amount in GBP: "))
                    self.deposit_to_customer(customer_id, amount)
                    print("Deposit successful.")

                elif choice == "3":
                    customer_id = int(input("Enter customer ID: "))
                    amount = float(input("Enter withdrawal amount in GBP: "))
                    self.withdraw_from_customer(customer_id, amount)
                    print("Withdrawal successful.")

                elif choice == "4":
                    from_id = int(input("Transfer from customer ID: "))
                    to_id = int(input("Transfer to customer ID: "))
                    amount = float(input("Enter amount in GBP: "))
                    self.transfer(from_id, to_id, amount)
                    print("Transfer successful.")

                elif choice == "5":
                    stats = self.transaction_statistics()
                    print("\nTransaction Statistics")
                    print("Count:", stats["count"])
                    print("Total: GBP", stats["total"])
                    print("Mean: GBP", stats["mean"])
                    print("Median: GBP", stats["median"])
                    print("Max: GBP", stats["max"])
                    print("Min: GBP", stats["min"])

                elif choice == "6":
                    self.generate_sample_data()
                    print("Sample data generated.")

                elif choice == "7":
                    self.save_customers_to_csv()
                    self.save_transactions_to_csv()
                    self.save_summary_to_txt()
                    print("Files saved.")

                elif choice == "8":
                    self.export_to_excel()
                    print("Excel report saved as bank_report.xlsx")

                elif choice == "9":
                    print("Next UK bank holiday:", self.get_next_bank_holiday())

                elif choice == "10":
                    print("Exiting program.")
                    break

                else:
                    print("Invalid choice. Please try again.")

            except ValueError as error:
                print("Error:", error)
            except Exception as error:
                print("Something went wrong:", error)

# Main function starts the whole program.
def main():
    bank = Bank()
    bank.menu()


# This makes sure the program runs only when this file is executed directly.
if __name__ == "__main__":
    main()